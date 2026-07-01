"""SharePoint REST API client.

Downloads files from SharePoint using the SharePoint REST API. Two
authentication methods are supported for obtaining the Bearer token:

* **Entra ID app-only with a certificate (recommended).** A certificate-signed
  client assertion is exchanged for a token via the Microsoft identity
  platform. This is Microsoft's replacement for the retired Azure ACS flow.
* **Legacy Azure ACS app-only (deprecated).** A client-credentials form body is
  POSTed to the Azure ACS token endpoint. Azure ACS for SharePoint Online was
  retired by Microsoft and stopped working on 2 April 2026.

In both cases the file is then fetched via the SharePoint REST API using the
resulting Bearer token — the REST calls are identical regardless of auth method.
"""

import base64
import csv
import io
import json
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from urllib.parse import quote, urlparse

import requests

from dlt_saga.pipelines.base_client import BaseClient
from dlt_saga.utility.optional_deps import require_optional
from dlt_saga.utility.secrets.resolver import resolve_secret

from .config import SharePointConfig

logger = logging.getLogger(__name__)

_TOKEN_ENDPOINT = (
    "https://accounts.accesscontrol.windows.net/{tenant_id}/tokens/OAuth/2"
)


class SharePointClient(BaseClient):
    """Client for downloading and parsing SharePoint files."""

    def __init__(self, config: SharePointConfig):
        self.config = config
        self._token: Optional[str] = None

    # ------------------------------------------------------------------
    # BaseClient interface
    # ------------------------------------------------------------------

    def connect(self, credentials=None):
        self._token = self._acquire_token()

    def test_connection(self):
        if not self._token:
            self.connect()
        logger.info("SharePoint token acquired — connection OK")

    # ------------------------------------------------------------------
    # Public helpers
    # ------------------------------------------------------------------

    def get_last_modified_time(self) -> datetime:
        """Return the last-modified timestamp of the SharePoint file (UTC)."""
        if not self._token:
            self.connect()

        encoded_path = quote(self.config.file_path, safe="/-_.")
        url = (
            f"{self.config.site_url.rstrip('/')}/"
            f"_api/web/GetFileByServerRelativeUrl('{encoded_path}')"
            f"?$select=TimeLastModified"
        )
        response = requests.get(
            url,
            headers={
                "Authorization": f"Bearer {self._token}",
                "Accept": "application/json;odata=nometadata",
            },
            timeout=30,
        )
        response.raise_for_status()
        raw = response.json().get("TimeLastModified")
        if not raw:
            raise ValueError(
                "SharePoint file metadata response did not contain 'TimeLastModified'"
            )
        dt = datetime.fromisoformat(raw.rstrip("Z")).replace(tzinfo=timezone.utc)
        logger.debug("SharePoint file last modified: %s", dt.isoformat())
        return dt

    def download_file(self) -> bytes:
        """Download a file from SharePoint and return raw bytes."""
        if not self._token:
            self.connect()

        encoded_path = quote(self.config.file_path, safe="/-_.")
        url = (
            f"{self.config.site_url.rstrip('/')}/"
            f"_api/web/GetFileByServerRelativeUrl('{encoded_path}')/$value"
        )
        logger.debug("Downloading SharePoint file: %s", url)
        response = requests.get(
            url,
            headers={"Authorization": f"Bearer {self._token}"},
            timeout=120,
        )
        response.raise_for_status()
        logger.info("Downloaded %s bytes from SharePoint", f"{len(response.content):,}")
        return response.content

    def read_file(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Dispatch to the correct reader based on config.file_type."""
        file_type = self.config.file_type.lower()
        if file_type == "xlsx":
            return self.read_excel(file_bytes)
        elif file_type == "csv":
            return self.read_csv(file_bytes)
        elif file_type == "json":
            return self.read_json(file_bytes)
        elif file_type == "jsonl":
            return self.read_jsonl(file_bytes)
        else:
            raise ValueError(f"Unsupported file_type: {file_type}")

    def read_excel(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Parse an Excel workbook and return rows as a list of dicts."""
        require_optional("openpyxl", "SharePoint Excel files")
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        try:
            if self.config.sheet_name:
                if self.config.sheet_name not in wb.sheetnames:
                    raise ValueError(
                        f"Sheet '{self.config.sheet_name}' not found. "
                        f"Available sheets: {wb.sheetnames}"
                    )
                ws = wb[self.config.sheet_name]
            else:
                ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            return []

        header_idx = self.config.header_row - 1
        headers = [
            str(cell) if cell is not None else f"col_{i}"
            for i, cell in enumerate(rows[header_idx])
        ]

        records: List[Dict[str, Any]] = []
        for row in rows[header_idx + 1 :]:
            if all(v is None for v in row):
                continue
            records.append(dict(zip(headers, row)))

        return records

    def read_csv(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Parse a CSV file and return rows as a list of dicts."""
        text = file_bytes.decode(self.config.encoding)
        reader = csv.DictReader(io.StringIO(text), delimiter=self.config.csv_separator)
        return [dict(row) for row in reader]

    def read_json(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Parse a JSON file (array at root) and return rows as a list of dicts."""
        data = json.loads(file_bytes.decode(self.config.encoding))
        if isinstance(data, list):
            return data
        raise ValueError(
            "JSON file must contain a root-level array. "
            f"Got {type(data).__name__} instead."
        )

    def read_jsonl(self, file_bytes: bytes) -> List[Dict[str, Any]]:
        """Parse a JSONL (newline-delimited JSON) file and return rows as a list of dicts."""
        text = file_bytes.decode(self.config.encoding)
        return [json.loads(line) for line in text.splitlines() if line.strip()]

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _acquire_token(self) -> str:
        """Acquire a SharePoint Bearer token via the configured auth method."""
        if self.config.certificate:
            return self._acquire_token_entra_id()
        if self.config.token_request_body:
            logger.warning(
                "SharePoint is authenticating via the legacy Azure ACS app-only "
                "flow ('token_request_body'). Azure ACS for SharePoint Online was "
                "retired by Microsoft and stopped working on 2 April 2026, so this "
                "path may fail. Migrate to Entra ID certificate authentication by "
                "setting 'client_id' and 'certificate'."
            )
            return self._acquire_token_acs()
        # Guarded by SharePointConfig._validate_auth(); defensive fallback.
        raise ValueError(
            "SharePoint authentication is not configured. Provide 'client_id' and "
            "'certificate', or the legacy 'token_request_body'."
        )

    def _acquire_token_entra_id(self) -> str:
        """Acquire a token via Entra ID app-only certificate authentication."""
        require_optional("azure.identity", "SharePoint certificate authentication")
        from azure.identity import CertificateCredential

        password = (
            resolve_secret(self.config.certificate_password).encode("utf-8")
            if self.config.certificate_password
            else None
        )
        credential = CertificateCredential(
            tenant_id=resolve_secret(self.config.tenant_id),
            client_id=resolve_secret(self.config.client_id),
            certificate_data=self._certificate_bytes(),
            password=password,
        )
        scope = f"{self._sharepoint_resource()}/.default"
        token = credential.get_token(scope)
        logger.debug(
            "SharePoint Entra ID token acquired (expires_on=%s)", token.expires_on
        )
        return token.token

    def _certificate_bytes(self) -> bytes:
        """Resolve the certificate secret to raw bytes for ``CertificateCredential``.

        Accepts either a PEM string (private key + certificate) or a
        base64-encoded PKCS#12 (PFX) blob. Azure Key Vault returns a
        certificate's backing secret as base64-encoded PKCS#12 when the
        certificate uses the default ``application/x-pkcs12`` content type, and
        as PEM text when it uses ``application/x-pem-file``.
        """
        raw = resolve_secret(self.config.certificate)
        if "-----BEGIN" in raw:
            return raw.encode("utf-8")
        try:
            # binascii.Error (raised on invalid base64) subclasses ValueError.
            return base64.b64decode(raw, validate=True)
        except ValueError as exc:
            raise ValueError(
                "SharePoint 'certificate' is neither PEM (missing '-----BEGIN') nor "
                "valid base64-encoded PKCS#12. Ensure the secret holds the "
                "certificate including its private key."
            ) from exc

    def _acquire_token_acs(self) -> str:
        """Acquire a token via the legacy Azure ACS app-only flow (deprecated)."""
        oauth_body = resolve_secret(self.config.token_request_body)
        url = _TOKEN_ENDPOINT.format(tenant_id=resolve_secret(self.config.tenant_id))
        response = requests.post(
            url,
            data=oauth_body,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()
        token = data.get("access_token")
        if not token:
            raise ValueError(
                "Token response did not contain 'access_token'. "
                f"Response keys: {list(data.keys())}"
            )
        logger.debug(
            "SharePoint Bearer token acquired (expires_in=%s)", data.get("expires_in")
        )
        return token

    def _sharepoint_resource(self) -> str:
        """Return the SharePoint resource root used to build the token scope.

        Derived from ``site_url`` — e.g. ``https://contoso.sharepoint.com/sites/X``
        yields ``https://contoso.sharepoint.com``.
        """
        parsed = urlparse(self.config.site_url)
        if not parsed.scheme or not parsed.netloc:
            raise ValueError(
                f"site_url must be an absolute URL, got '{self.config.site_url}'"
            )
        return f"{parsed.scheme}://{parsed.netloc}"
